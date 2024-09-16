# -*- coding: utf-8 -*-
"""
Created on August 2024
@author : Rachel Honnert TA74

Ce script rempli automatiquement le fichier FeuilleMensuelleCahierRS_2024.ods à partir des fichiers COR et REF 
Modifications : 
"""

# Chargement des librairies
import datetime
from openpyxl import load_workbook
import sys
import os
import codecs 
import numpy as np
import math
from math import sin, cos, sqrt, atan2, radians
############# DEFINITION DES FONCTIONS #############
#
def lecture_rs_cor(date,AltTropo):
  # Florent Tence : date sous forme DD-MM-AAAA
  # modifier par R. Honnert pour ajouter les vitesses ascensionnelles
  # modofier par R. Honnert pour sortir les valeurs marquantes des RS  
  jour = date[:2]
  mois = date[3:5]
  annee = date[6:10]

  dossier = sorted([f for f in os.listdir(emplacement_rs) if '.cor' in f]) # list tous les .cor du fichier
  radical_fichier_mf = 'DD' + annee + mois + jour # je recherche le fichier du jour en question 
  P_tot,alt_tot,T_tot,FF_tot,DD_tot,Lat_tot,Lon_tot = [], [], [], [], [], [], [] # je recupère l'altitude, la pression, la température, l'humidité, vent ff et direction, vitesse ascendente
  if [f for f in dossier if radical_fichier_mf in f] == []: # Si le dossier est vide, on affiche un message signalant qu'il n'y a pas de mesure pour cette date
    print("Il n'y a pas de mesure météo france pour cette date")
  else:
    # Ouverture et lecture du fichier de RS
    nom_fichier_meteofrance = [f for f in dossier if radical_fichier_mf in f][-1] # on prend le dernier lancer de la journée
    fichier_mf = codecs.open(emplacement_rs+'/'+nom_fichier_meteofrance,encoding = 'latin1')
    lines = fichier_mf.readlines()

    for i in range(1, len(lines)):
      tmp = lines[i].replace('\n', '').replace('\r','').replace(';', '\t').split('\t')
      if len(tmp) > 1:
        P_tot.append(float(tmp[12]))          # Pression atmosphérique hPa
        alt_tot.append(float(tmp[1]))
        Lat_tot.append(float(tmp[2]))   # Delta de latitude 
        Lon_tot.append(float(tmp[3]))   # Delta de longitude        
        T_tot.append(float(tmp[10]))   # Température en °C 
        FF_tot.append(float(tmp[7]))   # force du vent
        DD_tot.append(float(tmp[8]))   # direction du vent

    lat1=radians(-66.65)
    lon1=radians(140.00)
    if not not(alt_tot) :
      imax=alt_tot.index(max(alt_tot))
      Tmin=min(T_tot[0:imax])
      itropo=alt_tot[0:imax].index(min(alt_tot[0:imax], key=lambda x:abs(x-AltTropo)))
      PAltMax=round(P_tot[imax],1)
      TTropo=T_tot[itropo]
      FFVMax=round(max(FF_tot)/0.51444)
      ivmax=FF_tot.index(max(FF_tot))
      DDVMax=int(DD_tot[ivmax])
      AltVMax=round(alt_tot[ivmax])
      lat2=Lat_tot[imax]
      lon2=Lon_tot[imax]
      # Approximate radius of earth in km
      R = 6373.0
      dlon = lon2 - lon1
      dlat = lat2 - lat1
      a = sin(dlat / 2)**2 + cos(lat1) * cos(lat2) * sin(dlon / 2)**2
      c = 2 * atan2(sqrt(a), sqrt(1 - a))
      DO = round(R * c,1)
  return Tmin,TTropo,FFVMax,DDVMax,AltVMax,DO,PAltMax # 

def lecture_rs_ref(date):
  # récupération des données RS des fichiers REF
  # date sous forme DD-MM-AAAA => 
  jour = date[:2]
  mois = date[3:5]
  annee = date[6:10]
  filename=os.path.join(emplacement_rs,"DD"+annee+mois+jour+'00_1.ref')
  print(filename)
  with open(filename,'r') as f:
    for line in f:
      if line.startswith('SondeID'):
        tmp=line.rstrip('\n').split('=')
        SondeID=tmp[1]
      elif line.startswith('GroundP'):
        tmp=line.rstrip('\n').split('=')
        GroundP=float(tmp[1][:-3])
      elif line.startswith('GroundClouds'):
        tmp=line.rstrip('\n').split('=')
        GroundClouds=tmp[1]
      elif line.startswith('GroundWindSpeed'):
        tmp=line.rstrip('\n').split('=')
        GroundWindSpeed=float(tmp[1][:-3])
      elif line.startswith('GroundWindDir'):
        tmp=line.rstrip('\n').split('=')
        GroundWindDir=int(float(tmp[1][:-1]))
      elif line.startswith('GroundT'):
        tmp=line.rstrip('\n').split('=')
        GroundT=float(tmp[1][:-2])
      elif line.startswith('GroundU'):
        tmp=line.rstrip('\n').split('=')
        GroundU=int(float(tmp[1][:-1]))
      elif line.startswith('Tropo'):
        tmp=line.rstrip('\n').replace('='," ").split(" ")
        TropoAlt=int(float(tmp[1][:-1]))
      elif line.startswith('AltMax'):
        tmp=line.rstrip('\n').split('=')
        FinPTU=round(float(tmp[1][:-1]))
      elif line.startswith('Pressure'):
        tmp=line.rstrip('\n').split('=')
        FinPTUP=round(float(tmp[1][:-3]),1)
      elif line.startswith('SondeStart'):
        tmp=line.rstrip('\n').split('=')
        SondeStart=tmp[1][:-3]
      elif line.startswith('Ascent'):
        tmp=line.rstrip('\n').split('=')
        tmp2=tmp[1][:-3]
        Ascent=round(float(tmp2)/0.514444,1)

  return SondeStart, SondeID, GroundP, GroundClouds, GroundWindDir, GroundWindSpeed, GroundT, GroundU, TropoAlt, Ascent, FinPTU, FinPTUP

def insertion(tableur, data):
  # Noms des mois dans le fichier ODS
  mois_ods = ["Janvier","Février","Mars","Avril","Mai","Juin","Juillet","Août","Septembre","Octobre","Novembre","Décembre"]

  # Indices des colonnes
  indices = {}
  indices["heure"]  = 4
  indices["numero"] = 5
  indices["Psol"]   = 6
  indices["Nsol"]   = 7
  indices["DDsol"]  = 8
  indices["FFsol"]  = 9
  indices["Tsol"]   = 10
  indices["HUsol"]  = 11
  #indices["dateBallon"] = 12
  indices["Sol"]      = 13
  #indices["Cadre"]    = 14
  indices["Nbsonde"]  = 15
  indices["Nbballon"] = 16
  indices["NbDRVF"]   = 17
  indices["TropoT"]   = 18
  indices["TropoAlt"] = 19
  indices["Tmin"]     = 20
  indices["VentMaxDD"]= 21
  indices["VentMaxFF"]= 22
  indices["VentMaxalt"]= 23
  indices["DO"]       = 24
  indices["VA"]       = 25
  indices["FinPTU"]   = 28
  indices["hPa"]      = 29
  indices["Rem"]      = 30
  
  # Conversion du tableur ODS en XLSX
  #programme="libreoffice"
  programme = r'"C:/Program Files/LibreOffice/program/soffice.bin"'
  os.system(programme+" --headless --invisible --convert-to xlsx " + tableur) 

  # Lecture du tableur XLSX
  workbook = load_workbook(tableur.replace(".ods", ".xlsx"))

  # Récupération de la page du mois concerné
  page = workbook[mois_ods[data["date"].month-1] + " " + str(data["date"].year)]

  # Insertion des données
  for k,v in indices.items():
    # Récupération de la coordonnée dans le tableur
    coord = page.cell(row=data["date"].day+3, column=v).coordinate

    # Ajout des donnnées
    page[coord] = data[k]

    # Format réel ou entier
    if type(data[k]) is float:
      page[coord].number_format="0.0"
    elif type(data[k]) is int:
      page[coord].number_format="0"

  # Ecriture du tableur XLSX
  workbook.save(tableur.replace(".ods", ".xlsx"))

  # Conversion du tableur XLSX en ODS
  os.system(programme+" --headless --invisible --convert-to ods " + tableur.replace(".ods", ".xlsx"))

  # Suppression du tableur XLSX
  os.remove(tableur.replace(".ods", ".xlsx"))

#############  PROGRAMME PRINCIPAL #############

if (__name__ == "__main__"):
  # Nom du tableur
  emplacement_tableur="."
  tableur = os.path.join(emplacement_tableur,r"FeuilleMensuelleCahierRS_2024.ods")
  # Si le fichier est ouvert, le script ne fonctionne pas 
  # 
  dossier = [f for f in os.listdir(emplacement_tableur) if '.~lock.FeuilleMensuelleCahierRS_2024.ods' in f]
  if [f for f in dossier ] != []:
    print("Le fichier "+tableur+" est ouvert. Il faut le fermer !")
  else:
   if len(sys.argv)>1 :
     if int(float(sys.argv[1])) != 2024 :
       print("L'année "+" est fausse")
     else :
       dateini = datetime.datetime(int(float(sys.argv[1])),int(float(sys.argv[2])),int(float(sys.argv[3])))
   else:
     dateini = datetime.datetime.now()
   aujourdhui= datetime.datetime.now()
   date =  dateini.strftime("%d-%m-%Y")
   if aujourdhui < dateini.replace(hour=2, minute=2):
     # Il est plus tôt que 2UTC ou 12h locale
     # Le RS du jour n'est pas encore arrivé 
     print("Les données ne sont pas disponibles avant 12h locale")
   else:
     emplacement_rs=os.path.join("//mtoservice/Archives/Archives-Exploitation/Archives-Radiosondage/Archives_Data/",date[6:10])
     # Calcul des données => fichier REF
     SondeStart,SondeID,GroundP,GroundClouds,GroundWindDir,GroundWindSpeed,GroundT,GroundU,TropoAlt,Ascent,FinPTU,FinPTUP=lecture_rs_ref(date)
     # Calcul des données => fichier COR
     Tmin,TTropo,FFVMax,DDVMax,AltVMax,DO,PALtmax=lecture_rs_cor(date,TropoAlt)
     data = {}
     # Récupération des nouvelles données à insérer dans le tableur, sous forme de dictionnaire
     jour = int(date[0:2])
     mois = int(date[3:5])
     annee = int(date[6:10])
     data["date"] = datetime.date(annee, mois, jour)
     data["heure"]  = SondeStart
     data["numero"] = SondeID
     data["Psol"]   = GroundP 
     data["Nsol"]   = GroundClouds 
     data["DDsol"]  = GroundWindDir 
     data["FFsol"]  = GroundWindSpeed
     data["Tsol"]   = GroundT
     data["HUsol"]  = GroundU
     data["dateBallon"] = 13
     data["Sol"]      = '.'
     data["Cadre"]    = 15
     data["Nbsonde"]  = 1
     data["Nbballon"] = 1
     data["NbDRVF"]   = '.'
     data["TropoT"]   = TTropo
     data["TropoAlt"]= TropoAlt
     data["Tmin"]     = Tmin
     data["VentMaxDD"]= DDVMax
     data["VentMaxFF"]= FFVMax
     data["VentMaxalt"]= AltVMax
     data["DO"]       = DO
     data["VA"]       = Ascent
     data["FinPTU"]   = FinPTU
     data["hPa"]      = PALtmax
     data["Rem"]      = 'RAS'
     print(data)
     # Insertion des nouvelles données dans le tableur
     insertion(tableur, data)
